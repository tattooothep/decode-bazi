#!/usr/bin/env python3
"""Export approved Google Sheet text into Decode/Qimen i18n JSON bundles.

The script reads a local spreadsheet export (XLSX) so CI can run without
Google credentials. Source tabs and output paths are config-driven.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to read XLSX sheet exports") from exc


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG = ROOT / "scripts/sync/i18n-export.config.json"
DEFAULT_SOURCE = Path("/tmp/hourkey-sheet.xlsx")

LANGS = ("th", "en", "zh")
FIELDS = ("main", "short", "long", "advice", "warning")
PRODUCTION_STATUSES = {"approved", "locked"}
PRODUCTION_DECISIONS = {"allow"}
BLOCK_VALUES = {
    "draft_by_gpt",
    "needs_sinsae_review",
    "hold",
    "block",
    "block_until_health_spec_approved",
    "web_copied",
    "dev_written",
    "unknown",
}
SAFE_AUTHORITIES = {"gpt_original", "sinsae_original", "approved_original"}

ALIASES = {
    "code": ("text_code", "รหัสหลังบ้าน", "code", "backend_code"),
    "module_id": ("module_id", "§ หมวด", "module", "section"),
    "severity": ("severity",),
    "tone": ("tone",),
    "text_review_status": ("text_review_status", "review_status", "current_status", "สถานะ"),
    "production_text_decision": (
        "production_text_decision",
        "production_decision",
        "production_after_approve",
    ),
    "copyright_risk": ("copyright_risk",),
    "rewrite_required": ("rewrite_required",),
    "dev_edit_allowed": ("dev_edit_allowed",),
    "text_source_authority": ("text_source_authority", "source_authority"),
    "approved_by_sinsae": ("approved_by_sinsae",),
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_status(value: Any) -> str:
    return norm(value).lower().replace(" ", "_").replace("-", "_")


def norm_bool(value: Any, default: bool | None = None) -> bool | None:
    text = norm_status(value)
    if text in {"true", "yes", "y", "1", "checked", "allow"}:
        return True
    if text in {"false", "no", "n", "0", "unchecked", "deny"}:
        return False
    return default


def compact_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def checksum(data: Any) -> str:
    return hashlib.sha256(compact_json(data).encode("utf-8")).hexdigest()


def load_config(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as fh:
        return json.load(fh)


def header_map(ws: Any, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        key = norm(ws.cell(header_row, col).value)
        if key:
            headers[key] = col
    return headers


def get(row: dict[str, Any], key: str) -> Any:
    for alias in ALIASES.get(key, (key,)):
        if alias in row:
            return row[alias]
    return None


def is_qimen(module_id: str, code: str) -> bool:
    module = norm(module_id)
    module_key = norm_status(module)
    return module in {"§30", "30"} or module_key.startswith("§30") or module_key.startswith("qimen") or norm_status(code).startswith("qimen_")


def is_health(module_id: str, code: str) -> bool:
    module = norm(module_id)
    module_key = norm_status(module)
    return module in {"§19", "19"} or module_key.startswith("§19") or module_key.startswith("health") or norm_status(code).startswith("health_")


def row_to_entry(row: dict[str, Any], source_sheet: str) -> dict[str, Any]:
    code = norm(get(row, "code"))
    module_id = norm(get(row, "module_id"))
    entry: dict[str, Any] = {
        "module_id": module_id,
        "severity": norm(get(row, "severity")) or None,
        "tone": norm(get(row, "tone")) or None,
        "_source": {"sheet": source_sheet},
    }
    for lang in LANGS:
        entry[lang] = {}
        for field in FIELDS:
            value = norm(row.get(f"{field}_{lang}"))
            entry[lang][field] = value or None
    if not entry["module_id"]:
        entry["module_id"] = "§30" if is_qimen(module_id, code) else None
    return entry


def production_decision(row: dict[str, Any]) -> str:
    decision = norm_status(get(row, "production_text_decision"))
    if decision.startswith("allow"):
        return "allow_qimen_beta" if "qimen" in decision else "allow"
    if "block_until_health_spec_approved" in decision:
        return "block_until_health_spec_approved"
    if decision.startswith("hold"):
        return "hold"
    return decision


def production_errors(row: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    code = norm(get(row, "code"))
    module_id = norm(get(row, "module_id"))
    status = norm_status(get(row, "text_review_status"))
    decision = production_decision(row)
    authority = norm_status(get(row, "text_source_authority"))
    risk = norm_status(get(row, "copyright_risk"))
    rewrite_required = norm_bool(get(row, "rewrite_required"), default=None)
    dev_edit_allowed = norm_bool(get(row, "dev_edit_allowed"), default=None)

    values = {status, decision, authority, risk}
    if values & BLOCK_VALUES:
        errors.append("blocked_status_or_authority")
    if status not in PRODUCTION_STATUSES:
        errors.append("text_review_status_not_approved_or_locked")
    if decision not in PRODUCTION_DECISIONS:
        errors.append("production_text_decision_not_allow")
    if authority not in SAFE_AUTHORITIES:
        errors.append("unsafe_source_authority")
    if risk != "low":
        errors.append("copyright_risk_not_low")
    if rewrite_required is not False:
        errors.append("rewrite_required_not_false")
    if dev_edit_allowed is not False:
        errors.append("dev_edit_allowed_not_false")
    for lang in LANGS:
        if not norm(row.get(f"main_{lang}")):
            errors.append(f"missing_main_{lang}")
    if is_health(module_id, code):
        errors.append("health_19_blocked_from_production")
    return errors


def should_consider_for_production(row: dict[str, Any]) -> bool:
    decision = production_decision(row)
    status = norm_status(get(row, "text_review_status"))
    return status in PRODUCTION_STATUSES or decision in PRODUCTION_DECISIONS


def iter_rows(wb: Any, config: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    for sheet_cfg in config["source_sheets"]:
        name = sheet_cfg["name"]
        if name not in wb.sheetnames:
            if sheet_cfg.get("optional"):
                warnings.append(f"optional sheet missing: {name}")
                continue
            warnings.append(f"configured sheet missing: {name}")
            continue
        ws = wb[name]
        headers = header_map(ws, int(sheet_cfg.get("header_row", 1)))
        if not headers:
            warnings.append(f"sheet has no header: {name}")
            continue
        for row_idx in range(int(sheet_cfg.get("header_row", 1)) + 1, (ws.max_row or 0) + 1):
            row = {header: ws.cell(row_idx, col).value for header, col in headers.items()}
            code = norm(get(row, "code"))
            if not code:
                continue
            row["_source_sheet"] = name
            row["_source_row"] = row_idx
            rows.append(row)
    return rows, warnings


def empty_bundle(config: dict[str, Any], profile: str, review_mode: bool) -> dict[str, Any]:
    meta: dict[str, Any] = {
        "version": f"{config.get('version_prefix', 'decode-i18n')}.{dt.datetime.now(dt.timezone.utc).strftime('%Y%m%dT%H%M%SZ')}",
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source_sheet_id": config["spreadsheet_id"],
        "checksum": "",
        "export_profile": profile,
        "review_mode": review_mode,
        "entries_count": 0,
    }
    if review_mode:
        meta["warning"] = "review copy only; not production-approved"
    return {"_meta": meta, "entries": {}}


def build_bundles(rows: list[dict[str, Any]], config: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    bundles = {
        "decode.production": empty_bundle(config, "production", False),
        "decode.staging": empty_bundle(config, "staging", True),
        "qimen.production": empty_bundle(config, "production", False),
        "qimen.staging": empty_bundle(config, "staging", True),
    }
    errors: list[str] = []

    for row in rows:
        code = norm(get(row, "code"))
        module_id = norm(get(row, "module_id"))
        namespace = "qimen" if is_qimen(module_id, code) else "decode"
        entry = row_to_entry(row, norm(row.get("_source_sheet")))
        bundles[f"{namespace}.staging"]["entries"][code] = entry

        prod_errors = production_errors(row)
        if not prod_errors:
            bundles[f"{namespace}.production"]["entries"][code] = entry
            continue
        if should_consider_for_production(row):
            errors.append(f"{row.get('_source_sheet')}:{row.get('_source_row')}:{code}:{','.join(prod_errors)}")

    for bundle in bundles.values():
        entries = bundle["entries"]
        bundle["_meta"]["entries_count"] = len(entries)
        bundle["_meta"]["checksum"] = checksum(entries)
    return bundles, errors


def write_bundles(bundles: dict[str, Any], config: dict[str, Any], dry_run: bool) -> None:
    outputs = config["outputs"]
    mapping = {
        "decode.production": outputs["decode_production"],
        "decode.staging": outputs["decode_staging"],
        "qimen.production": outputs["qimen_production"],
        "qimen.staging": outputs["qimen_staging"],
    }
    for key, rel_path in mapping.items():
        path = ROOT / rel_path
        if dry_run:
            print(f"dry-run {rel_path}: {bundles[key]['_meta']['entries_count']} entries")
            continue
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as fh:
            json.dump(bundles[key], fh, ensure_ascii=False, indent=2)
            fh.write("\n")
        print(f"wrote {rel_path}: {bundles[key]['_meta']['entries_count']} entries")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(DEFAULT_SOURCE), help="Path to XLSX export from Google Sheets")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Export config JSON")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--allow-production-errors", action="store_true")
    args = parser.parse_args()

    config = load_config(Path(args.config))
    source = Path(args.source)
    if not source.exists():
        raise SystemExit(f"source workbook not found: {source}")

    wb = load_workbook(source, data_only=True)
    rows, warnings = iter_rows(wb, config)
    bundles, errors = build_bundles(rows, config)
    write_bundles(bundles, config, args.dry_run)

    for warning in warnings:
        print(f"warning: {warning}", file=sys.stderr)
    if errors and not args.allow_production_errors:
        print("production export validation failed:", file=sys.stderr)
        for error in errors[:50]:
            print(f"  {error}", file=sys.stderr)
        if len(errors) > 50:
            print(f"  ... {len(errors) - 50} more", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
