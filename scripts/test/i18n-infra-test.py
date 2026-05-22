#!/usr/bin/env python3
from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
EXPORT_SCRIPT = ROOT / "scripts/sync/sheet-to-i18n.py"


def load_export_module():
    spec = importlib.util.spec_from_file_location("sheet_to_i18n", EXPORT_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


def make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "07 · Frontend_Text 💬"
    headers = [
        "row_no",
        "module_id",
        "text_code",
        "main_th",
        "main_en",
        "main_zh",
        "severity",
        "tone",
        "text_review_status",
        "production_text_decision",
        "copyright_risk",
        "rewrite_required",
        "dev_edit_allowed",
        "text_source_authority",
    ]
    ws.append(["skip"])
    ws.append(["skip"])
    ws.append(["skip"])
    ws.append(headers)
    ws.append([1, "DAILY", "daily_ok", "TH", "EN", "ZH", "good", "premium", "approved", "ALLOW", "low", "FALSE", "FALSE", "GPT_original"])
    ws.append([2, "DAILY", "daily_draft", "TH", "EN", "ZH", "good", "premium", "draft_by_gpt", "HOLD", "low", "FALSE", "FALSE", "draft_by_gpt"])
    ws.append([3, "§30", "qimen_ok", "TH", "EN", "ZH", "good", "premium", "locked", "ALLOW", "low", "No", "0", "sinsae_original"])
    ws.append([4, "§19", "health_ok", "TH", "EN", "ZH", "good", "premium", "approved", "ALLOW", "low", "FALSE", "FALSE", "sinsae_original"])
    batch = wb.create_sheet("29 · GPT_Rewrite_Batch_09")
    batch.append(["skip"])
    batch.append(["skip"])
    batch.append(["skip"])
    batch.append(["row_no", "module_id", "text_code", "main_th", "main_en", "main_zh", "production_decision"])
    batch.append([1, "DAILY", "daily_batch_draft", "TH", "EN", "ZH", "HOLD_UNTIL_SINSAE_APPROVED"])
    wb.save(path)


def test_export_rules() -> None:
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        workbook = tmp / "sheet.xlsx"
        config = tmp / "config.json"
        out = tmp / "out"
        make_workbook(workbook)
        config.write_text(
            json.dumps(
                {
                    "spreadsheet_id": "10Spi6Xkeu9m7v1TDmHoTkyahU0g62NlgfI7vDqnAX_s",
                    "version_prefix": "test",
                    "source_sheets": [
                        {"name": "07 · Frontend_Text 💬", "header_row": 4},
                        {"name": "29 · GPT_Rewrite_Batch_09", "header_row": 4},
                    ],
                    "outputs": {
                        "decode_production": str((out / "decode.production.json").relative_to(ROOT)) if out.is_relative_to(ROOT) else str(out / "decode.production.json"),
                        "decode_staging": str(out / "decode.staging.json"),
                        "qimen_production": str(out / "qimen.production.json"),
                        "qimen_staging": str(out / "qimen.staging.json"),
                    },
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        module = load_export_module()
        wb_rows, _ = module.iter_rows(module.load_workbook(workbook, data_only=True), json.loads(config.read_text(encoding="utf-8")))
        bundles, errors = module.build_bundles(wb_rows, json.loads(config.read_text(encoding="utf-8")))

        assert "daily_ok" in bundles["decode.production"]["entries"]
        assert "daily_draft" not in bundles["decode.production"]["entries"]
        assert "daily_draft" in bundles["decode.staging"]["entries"]
        assert bundles["decode.staging"]["_meta"]["review_mode"] is True
        assert "warning" in bundles["decode.staging"]["_meta"]
        assert "qimen_ok" in bundles["qimen.production"]["entries"]
        assert "qimen_ok" not in bundles["decode.production"]["entries"]
        assert "health_ok" not in bundles["decode.production"]["entries"]
        assert any("health_19_blocked_from_production" in error for error in errors)


def test_write_bundles_skips_unchanged_entries() -> None:
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        module = load_export_module()
        entries = {"daily_ok": {"th": {"main": "TH"}}}
        bundle = {
            "_meta": {
                "version": "test.old",
                "generated_at": "old",
                "source_sheet_id": "sheet",
                "checksum": module.checksum(entries),
                "export_profile": "production",
                "review_mode": False,
                "entries_count": 1,
            },
            "entries": entries,
        }
        out_paths = {
            "decode_production": tmp / "decode.production.json",
            "decode_staging": tmp / "decode.staging.json",
            "qimen_production": tmp / "qimen.production.json",
            "qimen_staging": tmp / "qimen.staging.json",
        }
        for path in out_paths.values():
            path.write_text(json.dumps(bundle, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        new_bundle = {
            **bundle,
            "_meta": {
                **bundle["_meta"],
                "version": "test.new",
                "generated_at": "new",
            },
        }
        config = {"outputs": {key: str(path) for key, path in out_paths.items()}}
        module.write_bundles(
            {
                "decode.production": new_bundle,
                "decode.staging": new_bundle,
                "qimen.production": new_bundle,
                "qimen.staging": new_bundle,
            },
            config,
            False,
        )

        written = json.loads(out_paths["decode_production"].read_text(encoding="utf-8"))
        assert written["_meta"]["version"] == "test.old"


def test_wrapper_contract_static() -> None:
    src = (ROOT / "src/lib/i18n-decode.ts").read_text(encoding="utf-8")
    assert "export function t(" in src
    assert "[missing:i18n:${namespace}:${code}:${lang}:${field}]" in src
    assert "namespace: Namespace = \"decode\"" in src
    assert "qimenProduction" in src and "decodeProduction" in src


def test_daily_score_contract_static() -> None:
    src = (ROOT / "src/app/api/daily/score/route.ts").read_text(encoding="utf-8")
    assert "matched_codes" in src
    assert "text_codes" in src
    assert "raw_score" in src
    assert "weighted_score" in src
    assert "engine_version" in src
    assert "formula_id" in src
    assert "function labelOf" not in src
    assert "labelEn" not in src


def test_hardcoded_scanner_daily_score() -> None:
    result = subprocess.run(
        [
            sys.executable,
            str(ROOT / "scripts/sync/scan-hardcoded-text.py"),
            "--target",
            "src/app/api/daily/score/route.ts",
            "--fail-on-block",
            "--no-write",
        ],
        cwd=ROOT,
        text=True,
        capture_output=True,
    )
    assert result.returncode == 0, result.stderr + result.stdout


def main() -> int:
    for test in [
        test_export_rules,
        test_write_bundles_skips_unchanged_entries,
        test_wrapper_contract_static,
        test_daily_score_contract_static,
        test_hardcoded_scanner_daily_score,
    ]:
        test()
        print(f"ok {test.__name__}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
